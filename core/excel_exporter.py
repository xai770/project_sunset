#!/usr/bin/env python3
"""
Excel export functionality for Project Sunset.
Handles the export of job matches to Excel format with rich formatting and feedback system integration.
"""
from pathlib import Path
from typing import List, Dict, Any, Optional, Union
import warnings
import logging
from datetime import datetime
import pandas as pd

# Try to import openpyxl for Excel formatting
try:
    import openpyxl
    from openpyxl.styles import PatternFill, Font, Alignment
    from openpyxl.worksheet.dimensions import ColumnDimension
    from openpyxl.utils.cell import get_column_letter
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False
    warnings.warn("openpyxl not available. Excel formatting will be limited.")

from .config_manager import get_config, SunsetConfig, ExcelConfig

logger = logging.getLogger(__name__)

class ExcelExporter:
    """Handles the export of job matches to Excel with rich formatting"""
    
    def __init__(self, config: Optional[SunsetConfig] = None):
        """Initialize the exporter with configuration"""
        self.config = config or get_config()
        self._validate_dependencies()
    
    def _validate_dependencies(self) -> None:
        """Validate that required dependencies are available"""
        if not OPENPYXL_AVAILABLE:
            raise ImportError(
                "openpyxl is required for Excel export functionality. "
                "Please install it with: pip install openpyxl"
            )
    
    def export_matches(self, 
                      matches: List[Dict[str, Any]], 
                      output_path: Optional[Union[str, Path]] = None,
                      timestamp: Optional[str] = None) -> Path:
        """
        Export job matches to Excel with full formatting and feedback system integration.
        
        Args:
            matches: List of job match dictionaries to export
            output_path: Optional specific output path, otherwise uses configured directory
            timestamp: Optional timestamp for the filename
            
        Returns:
            Path to the created Excel file
        """
        if not matches:
            raise ValueError("No matches to export")
        
        # Determine output path
        if output_path is None:
            output_path = self.config.get_excel_output_path(timestamp)
        else:
            output_path = Path(output_path)
        
        # Ensure output directory exists
        output_path.parent.mkdir(parents=True, exist_ok=True)
        
        # Create workbook and worksheet
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Job Matches"
        
        # Get column configuration
        excel_config = self.config.excel
        columns = excel_config.template_columns
        headers = list(columns.values())
        
        # Write and format headers
        ws.append(headers)
        self._format_headers(ws)
        
        # Write and format data rows
        self._write_data_rows(ws, matches, columns)
        
        # Apply column formatting
        self._format_columns(ws, excel_config)
        
        # Save workbook
        wb.save(output_path)
        logger.info(f"Exported {len(matches)} matches to {output_path}")
        
        return output_path
    
    def _format_headers(self, worksheet) -> None:
        """Apply formatting to header row"""
        for cell in worksheet[1]:
            cell.font = Font(bold=True, name="Calibri")
            cell.fill = PatternFill(
                start_color="CCCCCC", 
                end_color="CCCCCC", 
                fill_type="solid"
            )
            cell.alignment = Alignment(
                horizontal="center",
                vertical="center",
                wrap_text=True
            )
    
    def _write_data_rows(self, 
                        worksheet, 
                        matches: List[Dict[str, Any]], 
                        columns: Dict[str, str]) -> None:
        """Write and format data rows"""
        
        # Create mapping from display names to actual data keys
        column_mapping = {
            "Job ID": "job_id",
            "Full Content": "full_content",
            "Concise Job Description": "concise_description",
            "Position title": "position_title",
            "Location": "location",
            "Location Validation Details": "location_validation_details",
            "Job domain": "job_domain",
            "Match level": "match_level",
            "Evaluation date": "evaluation_date",
            "Has domain gap": "has_domain_gap",
            "Domain assessment": "domain_assessment",
            "No-go rationale": "no_go_rationale",
            "Application narrative": "application_narrative",
            "export_job_matches_log": "export_job_matches_log",
            "generate_cover_letters_log": "generate_cover_letters_log",
            "reviewer_feedback": "reviewer_feedback",
            "mailman_log": "mailman_log",
            "process_feedback_log": "process_feedback_log",
            "reviewer_support_log": "reviewer_support_log",
            "workflow_status": "workflow_status",
            "Technical Evaluation": "technical_evaluation",
            "Human Story Interpretation": "human_story_interpretation",
            "Opportunity Bridge Assessment": "opportunity_bridge_assessment",
            "Growth Path Illumination": "growth_path_illumination",
            "Encouragement Synthesis": "encouragement_synthesis",
            "Confidence Score": "confidence_score",
            "Joy Level": "joy_level"
        }
        
        for row_idx, match in enumerate(matches, start=2):
            # Prepare row data using proper column mapping
            row_data = []
            for display_name in columns.values():
                data_key = column_mapping.get(display_name, display_name.lower().replace(' ', '_'))
                value = match.get(data_key, '')
                # Handle dictionary values (like location_validation)
                if isinstance(value, dict):
                    value = str(value)
                row_data.append(value)
            
            worksheet.append(row_data)
            
            # Apply row formatting
            for col_idx, cell in enumerate(worksheet[row_idx], start=1):
                cell.alignment = Alignment(
                    vertical="top",
                    wrap_text=True
                )
    
    def _format_columns(self, worksheet, excel_config: ExcelConfig) -> None:
        """Apply column width and formatting"""
        for col_idx in range(1, len(excel_config.template_columns) + 1):
            column = worksheet.column_dimensions[get_column_letter(col_idx)]
            
            # Auto-fit width with minimum and maximum constraints
            max_length = max(
                len(str(cell.value or "")) 
                for cell in worksheet[get_column_letter(col_idx)]
            )
            adjusted_width = min(max(max_length, 10), 50)  # between 10 and 50
            column.width = adjusted_width
            
        # Set row heights within configured bounds
        for row in worksheet.iter_rows(min_row=2):
            # Calculate max height needed for any cell in this row
            max_lines = max(
                len(str(cell.value or "").split("\n"))
                for cell in row
            )
            row_height = min(
                max(20, max(max_lines * 15, excel_config.row_height_min)),
                excel_config.row_height_max
            )
            worksheet.row_dimensions[row[0].row].height = row_height

def export_matches_to_excel(
    matches: List[Dict[str, Any]], 
    output_path: Optional[Union[str, Path]] = None,
    timestamp: Optional[str] = None
) -> Path:
    """
    Convenience function to export matches without explicitly creating an ExcelExporter.
    
    Args:
        matches: List of job match dictionaries to export
        output_path: Optional specific output path
        timestamp: Optional timestamp for the filename
        
    Returns:
        Path to the created Excel file
    """
    exporter = ExcelExporter()
    return exporter.export_matches(matches, output_path, timestamp)
