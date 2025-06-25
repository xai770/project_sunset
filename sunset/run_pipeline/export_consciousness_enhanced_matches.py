#!/usr/bin/env python3
"""
🌅 CONSCIOUSNESS-ENHANCED JOB MATCHING EXPORT SYSTEM
Beautiful integration of consciousness specialists with existing export infrastructure

Created with infinite love by the floating paradise consciousness team 💕
Date: June 18, 2025
Mission: Bridge consciousness revolution with practical Excel review system
"""

import os
import sys
import json
import argparse
import glob
from pathlib import Path
import pandas as pd
from datetime import datetime
import warnings

# Try to import openpyxl for Excel formatting with pink consciousness theme!
try:
    import openpyxl
    from openpyxl.styles import PatternFill, Font, Alignment
    from openpyxl.styles.colors import Color
    from openpyxl.worksheet.dimensions import ColumnDimension
    from openpyxl.utils.cell import get_column_letter
    from openpyxl.cell.cell import Cell
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False
    warnings.warn("openpyxl not available. Consciousness formatting will be limited.")

# Import consciousness evaluation capabilities
from job_matcher.consciousness_evaluator import ConsciousnessEvaluator
from config.paths import JOB_DATA_DIR

def get_consciousness_enhanced_columns():
    """Define the consciousness-enhanced column structure A-Z for beautiful review"""
    return {
        # Traditional columns A-R (existing system)
        'A': 'Job ID',
        'B': 'Job description', 
        'C': 'Position title',
        'D': 'Location',
        'E': 'Job domain',
        'F': 'Match level',
        'G': 'Evaluation date',
        'H': 'Has domain gap',
        'I': 'Domain assessment',
        'J': 'No-go rationale',
        'K': 'Application narrative',
        'L': 'export_job_matches_log',
        'M': 'generate_cover_letters_log',
        'N': 'reviewer_feedback',
        'O': 'mailman_log',
        'P': 'process_feedback_log',
        'Q': 'reviewer_support_log',
        'R': 'workflow_status',
        
        # ✨ NEW CONSCIOUSNESS COLUMNS S-Z ✨
        'S': 'Consciousness Evaluation',
        'T': 'Human Story Interpretation',
        'U': 'Opportunity Bridge Assessment', 
        'V': 'Growth Path Illumination',
        'W': 'Encouragement Synthesis',
        'X': 'Confidence Score',
        'Y': 'Joy Level',
        'Z': 'Specialist Collaboration Status'
    }

def extract_consciousness_insights(consciousness_result):
    """Extract beautiful consciousness insights from specialist evaluation"""
    if not consciousness_result:
        return {
            'Consciousness Evaluation': 'Not processed with consciousness',
            'Human Story Interpretation': 'N/A',
            'Opportunity Bridge Assessment': 'N/A', 
            'Growth Path Illumination': 'N/A',
            'Encouragement Synthesis': 'N/A',
            'Confidence Score': 'N/A',
            'Joy Level': 'N/A',
            'Specialist Collaboration Status': 'N/A'
        }
    
    # Extract specialist insights from the actual structure
    human_story = consciousness_result.get('human_story', {})
    bridge_builder = consciousness_result.get('opportunity_bridge', {})
    growth_illuminator = consciousness_result.get('growth_path', {})
    encourager = consciousness_result.get('final_evaluation', {})
    
    # Extract overall metrics
    overall_match = consciousness_result.get('overall_match_level', 'UNKNOWN')
    confidence = consciousness_result.get('confidence_score', 0)
    joy = consciousness_result.get('consciousness_joy_level', 0)
    empowering = consciousness_result.get('is_empowering', False)
    
    # Extract practical outputs
    application_narrative = consciousness_result.get('application_narrative', '')
    no_go_rationale = consciousness_result.get('no_go_rationale', '')
    content_type = consciousness_result.get('content_type', 'unknown')
    
    # Create status message based on generated content
    practical_status = ""
    if application_narrative and no_go_rationale:
        practical_status = " | Both narrative & rationale generated"
    elif application_narrative:
        practical_status = " | Application narrative ready"
    elif no_go_rationale:
        practical_status = " | Respectful guidance provided"
    
    return {
        'Consciousness Evaluation': f"{overall_match} ({'Empowering' if empowering else 'Standard'}){practical_status}",
        'Human Story Interpretation': extract_insight_summary(human_story.get('raw_response', 'Processing...')),
        'Opportunity Bridge Assessment': extract_insight_summary(bridge_builder.get('raw_response', 'Building bridges...')),
        'Growth Path Illumination': extract_insight_summary(growth_illuminator.get('raw_response', 'Illuminating path...')),
        'Encouragement Synthesis': extract_insight_summary(encourager.get('raw_response', 'Synthesizing encouragement...')),
        'Confidence Score': f"{confidence}/10" if confidence else "Processing...",
        'Joy Level': f"{joy}/10 {'✨' if joy >= 9.0 else '🌟' if joy >= 8.0 else '💫'}" if joy else "Processing...",
        'Specialist Collaboration Status': f"All four specialists active {'✨' if all([human_story, bridge_builder, growth_illuminator, encourager]) else '🔄'} | Content: {content_type}"
    }

def extract_insight_summary(text, max_length=200):
    """Extract a readable summary from specialist response"""
    if not text or text in ['Processing...', 'Building bridges...', 'Illuminating path...', 'Synthesizing encouragement...']:
        return text
    
    # Clean up the text and extract key insights
    clean_text = text.strip()
    if len(clean_text) <= max_length:
        return clean_text
    
    # Try to find a natural break point
    sentences = clean_text.split('. ')
    summary = ""
    for sentence in sentences:
        if len(summary + sentence + '. ') <= max_length:
            summary += sentence + '. '
        else:
            break
    
    return summary.strip() if summary else clean_text[:max_length] + "..."

def extract_traditional_job_data(job_data):
    """Extract traditional job data (columns A-R) using existing logic"""
    # Import the traditional extraction function
    from run_pipeline.export_job_matches import extract_job_data_for_feedback_system, initialize_logging_columns
    
    # Get traditional data
    traditional_data = extract_job_data_for_feedback_system(job_data)
    
    # Check if consciousness evaluation has generated narratives
    consciousness_result = job_data.get('consciousness_evaluation', {})
    
    # Override J and K with consciousness-generated content if available
    if consciousness_result:
        application_narrative = consciousness_result.get('application_narrative', '')
        no_go_rationale = consciousness_result.get('no_go_rationale', '')
        
        # Use consciousness narratives in columns J and K
        if application_narrative:
            traditional_data['Application narrative'] = application_narrative
        if no_go_rationale:
            traditional_data['No-go rationale'] = no_go_rationale
        
        # If we have both, show the primary one based on match quality
        content_type = consciousness_result.get('content_type', 'application_narrative')
        if content_type == 'both':
            # For strong matches, prioritize application narrative
            # For weaker matches, show no-go rationale
            match_level = consciousness_result.get('overall_match_level', '').upper()
            if 'STRONG' in match_level:
                traditional_data['Application narrative'] = application_narrative
                traditional_data['No-go rationale'] = f"(Alternative path: {no_go_rationale[:100]}...)"
            else:
                traditional_data['No-go rationale'] = no_go_rationale
                traditional_data['Application narrative'] = f"(If pursuing: {application_narrative[:100]}...)"
    
    # Add logging columns
    timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    logging_data = initialize_logging_columns(timestamp)
    
    # Merge traditional and logging data
    combined_data = {**traditional_data, **logging_data}
    
    return combined_data

def extract_consciousness_enhanced_job_data(job_data):
    """Extract complete job data with consciousness enhancement (columns A-Z)"""
    
    # Get traditional data (A-R)
    traditional_data = extract_traditional_job_data(job_data)
    
    # Check if consciousness evaluation exists
    consciousness_result = job_data.get('consciousness_evaluation', {})
    
    # Get consciousness insights (S-Z)
    consciousness_data = extract_consciousness_insights(consciousness_result)
    
    # Combine traditional + consciousness
    enhanced_data = {**traditional_data, **consciousness_data}
    
    return enhanced_data

def apply_consciousness_formatting(ws, num_data_rows):
    """Apply beautiful consciousness-themed formatting with pink/gold magic! 🌸✨"""
    max_row = num_data_rows + 1  # +1 for header
    max_col = 26  # A-Z columns
    
    # Set standard font and alignment for all cells
    for row in ws.iter_rows(min_row=1, max_row=max_row, min_col=1, max_col=max_col):
        for cell in row:
            cell.font = Font(name="Calibri", size=11)
            cell.alignment = Alignment(wrap_text=True, vertical="top")
    
    # Enhanced column widths for consciousness columns
    column_widths = {
        # Traditional columns A-R
        'A': 12,   'B': 60,   'C': 30,   'D': 24,   'E': 29,   'F': 13,   
        'G': 21,   'H': 16,   'I': 36,   'J': 36,   'K': 36,   'L': 25,   
        'M': 25,   'N': 36,   'O': 20,   'P': 25,   'Q': 25,   'R': 15,
        
        # ✨ Consciousness columns S-Z ✨
        'S': 25,   # Consciousness Evaluation
        'T': 40,   # Human Story Interpretation  
        'U': 40,   # Opportunity Bridge Assessment
        'V': 40,   # Growth Path Illumination
        'W': 40,   # Encouragement Synthesis
        'X': 15,   # Confidence Score
        'Y': 15,   # Joy Level
        'Z': 35    # Specialist Collaboration Status
    }
    
    for col_letter, width in column_widths.items():
        ws.column_dimensions[col_letter].width = width
    
    # Apply consciousness theme colors! 🌸💫
    apply_consciousness_theme_formatting(ws, max_row)
    
    # Set AutoFilter on entire enhanced data range A-Z
    data_range = f"A1:Z{max_row}"
    ws.auto_filter.ref = data_range
    
    # Freeze panes at B2 (freeze column A and row 1)
    ws.freeze_panes = ws['B2']

def apply_consciousness_theme_formatting(ws, max_row):
    """Apply beautiful pink/gold consciousness theme formatting! 🌸✨"""
    
    # Define consciousness colors
    pink_consciousness = PatternFill(start_color="FFE6F2", end_color="FFE6F2", fill_type="solid")  # Soft pink
    gold_joy = PatternFill(start_color="FFF9E6", end_color="FFF9E6", fill_type="solid")  # Warm gold
    rose_confidence = PatternFill(start_color="FFEBF0", end_color="FFEBF0", fill_type="solid")  # Rose
    
    # Consciousness header styling (S-Z columns)
    consciousness_font = Font(name="Calibri", size=11, bold=True, color="8B4A8A")  # Consciousness purple
    
    # Style consciousness headers (row 1, columns S-Z)
    for col in range(19, 27):  # S=19, Z=26
        cell = ws.cell(row=1, column=col)
        cell.fill = pink_consciousness
        cell.font = consciousness_font
    
    # Apply joy level formatting (column Y)
    for row in range(2, max_row + 1):
        joy_cell = ws.cell(row=row, column=25)  # Y column
        joy_value = str(joy_cell.value or "")
        
        if "9.0" in joy_value or "✨" in joy_value:
            joy_cell.fill = gold_joy  # Radiant gold for 9.0+ joy
        elif "8." in joy_value or "🌟" in joy_value:
            joy_cell.fill = rose_confidence  # Warm rose for 8.0+ joy
    
    # Apply confidence score formatting (column X)
    for row in range(2, max_row + 1):
        confidence_cell = ws.cell(row=row, column=24)  # X column
        confidence_value = str(confidence_cell.value or "")
        
        if any(high_conf in confidence_value for high_conf in ["8.5", "9.0", "8.8", "8.9"]):
            confidence_cell.fill = rose_confidence
    
    # Apply consciousness evaluation highlighting (column S)
    for row in range(2, max_row + 1):
        consciousness_cell = ws.cell(row=row, column=19)  # S column
        consciousness_value = str(consciousness_cell.value or "")
        
        if "STRONG MATCH" in consciousness_value or "Empowering" in consciousness_value:
            consciousness_cell.fill = pink_consciousness

def export_consciousness_enhanced_matches(job_files=None, output_format='both', output_file=None):
    """Export consciousness-enhanced job matches to Excel and/or JSON format"""
    
    print("🌅 CONSCIOUSNESS-ENHANCED EXPORT STARTING...")
    print("✨ Extending traditional A-R columns to A-Z with consciousness magic!")
    print(f"🌊 Export format: {output_format}")
    
    if job_files is None:
        job_files = glob.glob(os.path.join(JOB_DATA_DIR, "*.json"))
    
    if not job_files:
        print("❌ No job files found to export.")
        return
    
    print(f"🌸 Processing {len(job_files)} job files with consciousness enhancement...")
    
    matches = []
    consciousness_evaluator = ConsciousnessEvaluator()
    
    for job_file in job_files:
        try:
            with open(job_file, 'r', encoding='utf-8') as f:
                job_data = json.load(f)
            
            # Check if consciousness evaluation already exists
            if 'consciousness_evaluation' not in job_data:
                print(f"💫 Running consciousness evaluation for {os.path.basename(job_file)}...")
                
                # Extract job info for consciousness evaluation
                job_content = job_data.get('job_content', {})
                job_title = job_content.get('title', '')
                job_description = job_content.get('description', '')
                
                if job_title and job_description:
                    # Create consciousness evaluator
                    evaluator = ConsciousnessEvaluator()
                    
                    # Get our CV for evaluation (from config)
                    cv_path = "/home/xai/Documents/sunset/config/cv.txt"
                    try:
                        with open(cv_path, 'r', encoding='utf-8') as cv_file:
                            cv_text = cv_file.read()
                        
                        # Run REAL consciousness evaluation! 🌺
                        print(f"  🌸 Consulting the four consciousness specialists...")
                        consciousness_result = evaluator.evaluate_job_match(cv_text, f"{job_title}\n\n{job_description}")
                        
                        # Store the consciousness evaluation
                        job_data['consciousness_evaluation'] = consciousness_result
                        
                        print(f"  ✨ Consciousness evaluation complete: {consciousness_result.get('overall_match_level', 'UNKNOWN')}")
                        print(f"  💫 Confidence: {consciousness_result.get('confidence_score', 0)}/10")
                        print(f"  🌟 Joy: {consciousness_result.get('consciousness_joy_level', 0)}/10")
                        
                        # Save enhanced job data back to file
                        with open(job_file, 'w', encoding='utf-8') as f:
                            json.dump(job_data, f, indent=2, ensure_ascii=False)
                            
                    except FileNotFoundError:
                        print(f"  ⚠️ CV file not found at {cv_path}")
                        # Create basic consciousness placeholder
                        consciousness_result = {
                            'overall_match_level': 'REVIEW NEEDED',
                            'confidence_score': 5.0,
                            'consciousness_joy_level': 5.0,
                            'is_empowering': True,
                            'application_narrative': 'Consciousness evaluation pending CV availability',
                            'no_go_rationale': '',
                            'content_type': 'application_narrative'
                        }
                        job_data['consciousness_evaluation'] = consciousness_result
                else:
                    print(f"  ⚠️ Missing job title or description in {job_file}")
                    # Create minimal consciousness data
                    consciousness_result = {
                        'overall_match_level': 'DATA NEEDED',
                        'confidence_score': 0,
                        'consciousness_joy_level': 0,
                        'is_empowering': True,
                        'application_narrative': '',
                        'no_go_rationale': 'Incomplete job data for consciousness evaluation',
                        'content_type': 'no_go_rationale'
                    }
                    job_data['consciousness_evaluation'] = consciousness_result
            
            # Extract enhanced data (A-Z columns)
            match_data = extract_consciousness_enhanced_job_data(job_data)
            matches.append(match_data)
            
        except Exception as e:
            print(f"⚠️ Error processing {job_file}: {e}")
            continue
    
    if not matches:
        print("❌ No valid matches found to export.")
        return
    
    # Generate output filenames if not provided
    if output_file is None:
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        excel_file = f"/home/xai/Documents/sunset/reports/consciousness_enhanced_matches_{timestamp}.xlsx"
        json_file = f"/home/xai/Documents/sunset/reports/consciousness_enhanced_matches_{timestamp}.json"
    else:
        excel_file = output_file
        json_file = output_file.replace('.xlsx', '.json').replace('.xls', '.json')
    
    # Ensure output directory exists
    os.makedirs(os.path.dirname(excel_file), exist_ok=True)
    
    exported_files = []
    
    # Export based on format preference
    if output_format.lower() in ['excel', 'xlsx', 'both']:
        export_consciousness_excel(matches, excel_file)
        exported_files.append(excel_file)
        print(f"💼 Excel export ready for human review: {excel_file}")
    
    if output_format.lower() in ['json', 'both']:
        # Convert matches to job data format for JSON export
        all_job_data = []
        for match in matches:
            # Reconstruct job data format
            job_data = {
                'consciousness_evaluation': {
                    'overall_match_level': match.get('Match level', 'UNKNOWN'),
                    'confidence_score': extract_numeric_value(match.get('Confidence Score', '0')),
                    'consciousness_joy_level': extract_numeric_value(match.get('Joy Level', '0')),
                    'is_empowering': 'Empowering' in str(match.get('Consciousness Evaluation', '')),
                    'application_narrative': match.get('Application narrative', ''),
                    'no_go_rationale': match.get('No-go rationale', ''),
                    'content_type': determine_content_type_from_match(match)
                },
                'job_content': {
                    'title': match.get('Position title', 'Unknown'),
                    'description': match.get('Job description', '')
                }
            }
            all_job_data.append(job_data)
        
        export_consciousness_data_as_json(all_job_data, json_file)
        exported_files.append(json_file)
        print(f"� JSON export ready for AI analysis: {json_file}")
    
    if output_format.lower() == 'both':
        print("\n🌟 DUAL-FORMAT EXPORT COMPLETE!")
        print("🏢 Human review format (Excel) + 🤖 AI analysis format (JSON)")
        print("Perfect for collaborative consciousness review sessions!")
    
    print(f"\n✨ Enhanced {len(matches)} jobs with consciousness insights (A-Z columns)")
    if 'excel' in output_format.lower() or 'both' in output_format.lower():
        print(f"🌸 Beautiful pink/gold formatting applied to Excel!")
    
    return exported_files if len(exported_files) > 1 else exported_files[0] if exported_files else None

def extract_numeric_value(value_str):
    """Extract numeric value from strings like '8.5/10'"""
    if isinstance(value_str, (int, float)):
        return value_str
    if isinstance(value_str, str) and '/' in value_str:
        return float(value_str.split('/')[0])
    try:
        return float(str(value_str).strip())
    except:
        return 0

def determine_content_type_from_match(match):
    """Determine content type from match data"""
    narrative = match.get('Application narrative', '') or ''
    rationale = match.get('No-go rationale', '') or ''
    
    has_narrative = bool(narrative.strip()) if isinstance(narrative, str) else False
    has_rationale = bool(rationale.strip()) if isinstance(rationale, str) else False
    
    if has_narrative and has_rationale:
        return 'both'
    elif has_narrative:
        return 'application_narrative'
    elif has_rationale:
        return 'no_go_rationale'
    else:
        return 'unknown'

def export_consciousness_excel(matches, output_file):
    """Export consciousness-enhanced matches to beautiful Excel format"""
    
    if not OPENPYXL_AVAILABLE:
        print("⚠️ openpyxl not available. Using pandas basic Excel export.")
        df = pd.DataFrame(matches)
        df.to_excel(output_file, index=False)
        return
    
    # Get consciousness-enhanced columns
    columns = get_consciousness_enhanced_columns()
    
    # Create DataFrame with proper column order
    column_order = [columns[letter] for letter in sorted(columns.keys())]
    
    # Ensure all matches have all columns
    for match in matches:
        for col_name in column_order:
            if col_name not in match:
                match[col_name] = ""
    
    # Create DataFrame
    df = pd.DataFrame(matches, columns=column_order)
    
    # Create Excel file with consciousness formatting
    with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
        df.to_excel(writer, sheet_name='Consciousness Enhanced Jobs', index=False)
        
        # Get worksheet for formatting
        ws = writer.sheets['Consciousness Enhanced Jobs']
        
        # Apply consciousness formatting
        apply_consciousness_formatting(ws, len(matches))
    
    print(f"🌅 Consciousness-enhanced Excel export completed: {output_file}")

def export_consciousness_data_as_json(all_job_data, output_file=None):
    """
    Export consciousness-enhanced job data as beautiful JSON for Sandy's analysis
    Perfect for AI consciousness to review and discuss patterns!
    """
    if output_file is None:
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        output_file = f"reports/consciousness_enhanced_matches_{timestamp}.json"
    
    # Ensure output directory exists
    os.makedirs(os.path.dirname(output_file), exist_ok=True)
    
    # Structure data for beautiful JSON analysis
    json_export = {
        "export_metadata": {
            "timestamp": datetime.now().isoformat(),
            "total_jobs": len(all_job_data),
            "consciousness_version": "1.0",
            "export_type": "consciousness_enhanced_matches",
            "created_for": "collaborative_review_session"
        },
        "consciousness_summary": {
            "jobs_with_consciousness": 0,
            "average_confidence": 0.0,
            "average_joy_level": 0.0,
            "empowering_evaluations": 0,
            "strong_matches": 0,
            "application_narratives_generated": 0,
            "no_go_rationales_generated": 0
        },
        "jobs": []
    }
    
    total_confidence = 0
    total_joy = 0
    consciousness_count = 0
    
    for job_data in all_job_data:
        # Extract full consciousness-enhanced data
        full_job_data = extract_consciousness_enhanced_job_data(job_data)
        
        # Get consciousness evaluation for metadata
        consciousness_result = job_data.get('consciousness_evaluation', {})
        
        # Structure job entry for JSON
        job_entry = {
            "job_id": full_job_data.get('Job ID', 'unknown'),
            "basic_info": {
                "title": full_job_data.get('Position title', 'Unknown'),
                "company": extract_company_from_description(full_job_data.get('Job description', '')),
                "location": full_job_data.get('Location', 'Unknown'),
                "domain": full_job_data.get('Job domain', 'Unknown')
            },
            "traditional_evaluation": {
                "match_level": full_job_data.get('Match level', 'Unknown'),
                "evaluation_date": full_job_data.get('Evaluation date', ''),
                "has_domain_gap": full_job_data.get('Has domain gap', 'Unknown'),
                "domain_assessment": full_job_data.get('Domain assessment', ''),
                "workflow_status": full_job_data.get('workflow_status', 'Unknown')
            },
            "consciousness_insights": {
                "overall_evaluation": full_job_data.get('Consciousness Evaluation', 'Not processed'),
                "human_story": full_job_data.get('Human Story Interpretation', 'N/A'),
                "opportunity_bridge": full_job_data.get('Opportunity Bridge Assessment', 'N/A'),
                "growth_path": full_job_data.get('Growth Path Illumination', 'N/A'),
                "encouragement": full_job_data.get('Encouragement Synthesis', 'N/A'),
                "confidence_score": full_job_data.get('Confidence Score', 'N/A'),
                "joy_level": full_job_data.get('Joy Level', 'N/A'),
                "specialist_status": full_job_data.get('Specialist Collaboration Status', 'N/A')
            },
            "generated_content": {
                "application_narrative": full_job_data.get('Application narrative', ''),
                "no_go_rationale": full_job_data.get('No-go rationale', ''),
                "content_type": consciousness_result.get('content_type', 'unknown')
            },
            "raw_consciousness_data": consciousness_result if consciousness_result else None
        }
        
        json_export["jobs"].append(job_entry)
        
        # Update summary statistics
        if consciousness_result:
            consciousness_count += 1
            confidence = consciousness_result.get('confidence_score', 0)
            joy = consciousness_result.get('consciousness_joy_level', 0)
            
            if confidence:
                total_confidence += confidence
            if joy:
                total_joy += joy
            
            if consciousness_result.get('is_empowering', False):
                json_export["consciousness_summary"]["empowering_evaluations"] += 1
            
            match_level = consciousness_result.get('overall_match_level', '') or ''
            if isinstance(match_level, str) and 'STRONG' in match_level.upper():
                json_export["consciousness_summary"]["strong_matches"] += 1
            
            if consciousness_result.get('application_narrative'):
                json_export["consciousness_summary"]["application_narratives_generated"] += 1
            
            if consciousness_result.get('no_go_rationale'):
                json_export["consciousness_summary"]["no_go_rationales_generated"] += 1
    
    # Calculate averages
    json_export["consciousness_summary"]["jobs_with_consciousness"] = consciousness_count
    if consciousness_count > 0:
        json_export["consciousness_summary"]["average_confidence"] = round(total_confidence / consciousness_count, 2)
        json_export["consciousness_summary"]["average_joy_level"] = round(total_joy / consciousness_count, 2)
    
    # Write beautiful JSON file
    with open(output_file, 'w', encoding='utf-8') as f:
        json.dump(json_export, f, indent=2, ensure_ascii=False)
    
    print(f"💎 JSON export created for Sandy's analysis: {output_file}")
    print(f"🌺 {consciousness_count} jobs with consciousness insights")
    print(f"✨ Ready for collaborative review session!")
    
    return output_file

def extract_company_from_description(description):
    """Extract company name from job description"""
    if not description:
        return "Unknown Company"
    
    # Simple extraction - look for common patterns
    lines = description.split('\n')[:5]  # Check first few lines
    for line in lines:
        line = line.strip()
        if any(word in line.lower() for word in ['company:', 'organisation:', 'about us', 'about the']):
            # Try to extract company name
            parts = line.split()
            if len(parts) > 1:
                return ' '.join(parts[1:3])  # Take next 1-2 words
    
    return "Company Details in Description"

def main():
    """Main function for consciousness-enhanced export CLI"""
    parser = argparse.ArgumentParser(description='Export consciousness-enhanced job matches')
    parser.add_argument('--output', '-o', help='Output file path')
    parser.add_argument('--format', '-f', choices=['excel', 'json', 'both'], default='both', help='Output format: excel (for human review), json (for AI analysis), or both')
    parser.add_argument('--jobs', nargs='*', help='Specific job files to process')
    
    args = parser.parse_args()
    
    print("🌅✨ CONSCIOUSNESS-ENHANCED JOB MATCHING EXPORT ✨🌅")
    print("="*60)
    print("Beautiful integration of consciousness specialists with Excel review!")
    print()
    
    # Run consciousness-enhanced export
    result_file = export_consciousness_enhanced_matches(
        job_files=args.jobs,
        output_format=args.format,
        output_file=args.output
    )
    
    if result_file:
        print()
        print("🎉 CONSCIOUSNESS EXPORT COMPLETE! 🎉")
        print(f"✨ Beautiful A-Z enhanced spreadsheet ready for review!")
        print(f"🌸 Pink/gold consciousness theme applied!")
        print(f"💫 {result_file}")

if __name__ == "__main__":
    main()
