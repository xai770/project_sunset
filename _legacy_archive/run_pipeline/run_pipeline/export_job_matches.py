#!/usr/bin/env python3
"""
Export job matching results to CSV or Excel format.

This script reads job JSON files from the data/postings directory and exports
the matching results to a table format (CSV or Excel) for easier review and analysis.
It specifically focuses on extracting domain knowledge assessment information.
"""
import os
import sys
import json
import argparse
import glob
from pathlib import Path
import pandas as pd
from datetime import datetime
import warnings

# Try to import openpyxl for Excel formatting
try:
    import openpyxl
    from openpyxl.styles import PatternFill, Font, Alignment
    from openpyxl.styles.colors import Color
    from openpyxl.worksheet.dimensions import ColumnDimension
    from openpyxl.utils.cell import get_column_letter
    from openpyxl.cell.cell import Cell
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False
    warnings.warn("openpyxl not available. Excel formatting will be limited.")

try:
    from run_pipeline.config.paths import JOB_DATA_DIR
except ImportError:
    # Fallback for direct or external import
    import sys, os
    from pathlib import Path
    PROJECT_ROOT = Path(os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__)))))
    JOB_DATA_DIR = PROJECT_ROOT / "data" / "postings"

def get_standard_columns():
    """Define the standard column structure A-R for feedback system"""
    return {
        'A': 'Job ID',
        'B': 'Job description',
        'C': 'Position title',
        'D': 'Location',
        'E': 'Job domain',
        'F': 'Match level',
        'G': 'Evaluation date',
        'H': 'Has domain gap',
        'I': 'Domain assessment',
        'J': 'No-go rationale',
        'K': 'Application narrative',
        'L': 'export_job_matches_log',
        'M': 'generate_cover_letters_log',
        'N': 'reviewer_feedback',
        'O': 'mailman_log',
        'P': 'process_feedback_log',
        'Q': 'reviewer_support_log',
        'R': 'workflow_status'
    }

def extract_job_domain_from_data(job_data):
    """Extract job domain from job data"""
    # Try multiple fields for domain information
    web_details = job_data.get("web_details", {})
    search_details = job_data.get("search_details", {})
    llama_eval = job_data.get("llama32_evaluation", {})
    
    # Check various possible domain fields
    domain = (
        job_data.get('domain') or 
        web_details.get('domain') or 
        search_details.get('JobCategory', [{}])[0].get('Name', '') if search_details.get('JobCategory') else ''
    )
    
    # If domain is still empty, try to extract it from domain knowledge assessment
    if not domain and llama_eval:
        domain_assessment = llama_eval.get('domain_knowledge_assessment', '')
        if domain_assessment:
            # Try to extract domain from first sentence of assessment
            first_sentence = domain_assessment.split('.')[0]
            if 'domain' in first_sentence.lower() or 'industry' in first_sentence.lower():
                domain = first_sentence
    
    # If we still don't have a domain, use job title to infer domain
    if not domain:
        job_title = web_details.get('position_title', '')
        
        # Try to infer domain from job title
        from run_pipeline.job_matcher.domain_analyzer import extract_job_domain
        if job_title:
            domain = extract_job_domain(job_title)
    
    # Default value if still empty
    if not domain:
        domain = "Unclassified"
        
    return domain


def extract_job_data_for_feedback_system(job_data):
    """Extract job data in the A-R column format for feedback system using beautiful JSON structure"""
    # Extract from beautiful JSON structure
    job_metadata = job_data.get("job_metadata", {})
    job_content = job_data.get("job_content", {})
    
    # Check both evaluation fields (new llama32_evaluation and legacy evaluation_results)
    llama_evaluation = job_data.get("llama32_evaluation", {})
    evaluation_results = job_data.get("evaluation_results", {})
    
    # Use llama32_evaluation as primary source, fallback to evaluation_results
    evaluation_data = llama_evaluation if llama_evaluation else evaluation_results
    
    processing_log = job_data.get("processing_log", [])
    
    # Get job_id from metadata
    job_id = job_metadata.get('job_id', '')
    
    # Extract location from beautiful structure
    location_data = job_content.get("location", {})
    city = location_data.get("city", "")
    country = location_data.get("country", "")
    location = f"{city}, {country}".strip(', ')
    
    # Extract evaluation data from llama32_evaluation or evaluation_results
    eval_date = evaluation_data.get('evaluation_date', '')
    match_level = evaluation_data.get('cv_to_role_match', '')
    domain_assessment = evaluation_data.get('domain_knowledge_assessment', '')
    
    # Extract no-go rationale with improved logic to handle consciousness evaluation
    no_go_rationale = ''
    application_narrative = ''
    
    # First, try to get from consciousness evaluation fields
    if evaluation_data.get('no_go_rationale'):
        no_go_rationale = evaluation_data.get('no_go_rationale', '')
    if evaluation_data.get('application_narrative'):
        application_narrative = evaluation_data.get('application_narrative', '')
    
    # Fallback to traditional evaluation structure
    if not no_go_rationale and not application_narrative:
        no_go_rationale = evaluation_data.get('decision', {}).get('rationale', '')
        application_narrative = evaluation_data.get('Application narrative', '')
    
    # Use job description from beautiful structure
    job_description = job_content.get('description', '')
    if not job_description:
        # Fallback to creating description from title and location
        title = job_content.get('title', '')
        job_description = f"Position: {title}" if title else ''
        if location:
            job_description += f" | Location: {location}"
    
    # Extract job domain from organization division
    job_domain = job_content.get('organization', {}).get('division', '') or 'Unclassified'
    
    # Set application narrative based on match level with improved logic
    if not application_narrative:
        if match_level == "Good":
            application_narrative = "Application narrative was not provided for this Good match"
        else:
            application_narrative = "N/A - Not a Good match"
    
    # Clean up no-go rationale to remove malformed formatting
    if no_go_rationale:
        # Remove the problematic wrapping text from malformed rationales
        if "Extracted from incorrectly formatted narrative:" in no_go_rationale:
            # Extract the actual content and reformat properly
            import re
            match = re.search(r'\[Extracted from incorrectly formatted narrative: (.*?)\]', no_go_rationale, re.DOTALL)
            if match:
                extracted_content = match.group(1).strip()
                # If the extracted content is positive (contains "I believe", "confident", "experience"), 
                # create a proper no-go rationale
                if any(positive_word in extracted_content.lower() for positive_word in ['believe', 'confident', 'experience', 'bring', 'contribute']):
                    no_go_rationale = "After careful consideration of my background and the role requirements, I have decided not to apply for this position at this time."
                else:
                    no_go_rationale = f"I have compared my CV and the role description and decided not to apply due to: {extracted_content}"
        
        # If still empty or too generic, provide a proper fallback
        if not no_go_rationale.strip() or "but no specific reasons were provided" in no_go_rationale:
            if match_level == "Low":
                no_go_rationale = "After careful consideration, I have decided this role may not be the best fit for my current experience and career goals."
            elif match_level == "Moderate":
                no_go_rationale = "While I appreciate the opportunity, I feel there may be a better alignment between my experience and other positions."
            else:
                no_go_rationale = ""  # Good matches shouldn't have no-go rationales
    
    return {
        'Job ID': job_id,
        'Job description': job_description,
        'Position title': job_content.get('title', ''),
        'Location': location,
        'Job domain': job_domain,
        'Match level': match_level,
        'Evaluation date': eval_date,
        'Has domain gap': determine_domain_gap_from_beautiful_format(evaluation_data),
        'Domain assessment': domain_assessment,
        'No-go rationale': no_go_rationale,
        'Application narrative': application_narrative
    }

def determine_domain_gap_from_beautiful_format(evaluation_results):
    """Determine if there's a domain gap based on beautiful format evaluation"""
    if not evaluation_results:
        return 'Unknown'
    
    # Look for domain gap indicators in the evaluation
    domain_assessment = (evaluation_results.get('domain_knowledge_assessment') or '').lower()
    decision_rationale = (evaluation_results.get('decision', {}).get('rationale') or '').lower()
    
    # Check for gap indicators
    gap_indicators = ['lacks', 'missing', 'no experience', 'gap', 'would take', 'years to acquire']
    has_gap = any(indicator in domain_assessment or indicator in decision_rationale 
                  for indicator in gap_indicators)
    
    return 'Yes' if has_gap else 'No'

def initialize_logging_columns(timestamp):
    """Initialize the logging columns L-R with default values"""
    return {
        'export_job_matches_log': f"Exported at {timestamp}, v1.0, success",
        'generate_cover_letters_log': 'Pending cover letter generation',
        'reviewer_feedback': '',
        'mailman_log': '',
        'process_feedback_log': '',
        'reviewer_support_log': '',
        'workflow_status': 'Exported'
    }

def export_to_csv(matches, output_file):
    """Export matches to CSV format (legacy function)"""
    df = pd.DataFrame(matches)
    df.to_csv(output_file, index=False, quoting=1)  # QUOTE_ALL

def export_to_excel(matches, output_file):
    """Export matches to Excel format (legacy function)"""
    if not matches:
        raise ValueError("No matches to export.")
    
    first = matches[0]
    if isinstance(first, dict):
        header = list(first.keys())  # Convert dict_keys to list
        rows = [list(m.values()) for m in matches]  # Convert dict_values to list
    elif isinstance(first, list):
        header = [f"col_{i+1}" for i in range(len(first))]
        rows = matches
    else:
        raise TypeError("Matches must be a list of dicts or list of lists.")
    
    df = pd.DataFrame(rows, columns=header)
    df.to_excel(output_file, index=False)
    return output_file

def calculate_row_height(match_data, columns):
    """Calculate appropriate row height based on content length"""
    max_lines = 1
    
    # Define column widths for line calculation
    column_widths = {
        'URL': 12, 'Job description': 60, 'Position title': 30, 'Location': 24,
        'Job domain': 29, 'Match level': 10,'Evaluation date': 21, 'Has domain gap': 16,
        'Domain assessment': 36, 'No-go rationale': 36, 'Application narrative': 36,
        'export_job_matches_log': 25, 'generate_cover_letters_log': 25, 'reviewer_feedback': 36,
        'mailman_log': 20, 'process_feedback_log': 25, 'reviewer_support_log': 25, 'workflow_status': 15
    }
    
    # Check each column for content that might wrap
    for col_name in columns.values():
        content = str(match_data.get(col_name, ''))
        if not content:
            continue
            
        col_width = column_widths.get(col_name, 20)
        
        # Estimate characters per line (roughly 1.2 chars per width unit)
        chars_per_line = int(col_width * 1.2)
        
        # Count explicit line breaks
        explicit_lines = content.count('\n') + 1
        
        # Calculate wrapped lines for long content
        if len(content) > chars_per_line:
            wrapped_lines = (len(content) // chars_per_line) + 1
        else:
            wrapped_lines = 1
        
        # Take the maximum of explicit and wrapped lines
        total_lines = max(explicit_lines, wrapped_lines)
        max_lines = max(max_lines, total_lines)
    
    # Convert lines to row height (roughly 15pt per line, minimum 30pt, maximum 200pt)
    row_height = max(30, min(200, max_lines * 15))
    return row_height

def export_to_excel_feedback_system(matches, output_file):
    """Export matches to Excel with full JMFS feedback system formatting"""
    wb = openpyxl.Workbook()
    ws = wb.active
    if ws is not None:  # Type guard
        ws.title = "Job Matches"
    else:
        raise RuntimeError("Failed to create Excel worksheet")
    
    # Get column structure
    columns = get_standard_columns()
    headers = list(columns.values())
    
    # Write headers
    ws.append(headers)
    
    # Format header row
    for cell in ws[1]:
        cell.font = Font(bold=True, name="Calibri")
        cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center")
    
    # Write data rows
    for row_idx, match in enumerate(matches, start=2):
        row_data = [match.get(col_name, '') for col_name in columns.values()]
        ws.append(row_data)
        
        # Set hyperlink in column A (Job ID)
        job_id = match.get('Job ID', '')
        cell_a = ws.cell(row=row_idx, column=1)
        
        if job_id and cell_a is not None:
            # Create hyperlink to Deutsche Bank careers page with correct URL format
            url = f"https://careers.db.com/professionals/search-roles/#/professional/job/{job_id}"
            display_text = f"Job {job_id}"
            cell_a.value = display_text  # type: ignore
            cell_a.hyperlink = url  # type: ignore
            cell_a.style = "Hyperlink"  # type: ignore
        elif cell_a is not None:
            cell_a.value = job_id  # type: ignore
        
        # Calculate and set dynamic row height based on content
        row_height = calculate_row_height(match, columns)
        ws.row_dimensions[row_idx].height = row_height
    
    # Apply formatting
    apply_feedback_system_formatting(ws, len(matches))
    
    # Save workbook
    wb.save(output_file)
    print(f"Excel file exported with {len(matches)} jobs to: {output_file}")

def apply_feedback_system_formatting(ws, num_data_rows):
    """Apply comprehensive formatting to match the old working Excel format"""
    max_row = num_data_rows + 1  # +1 for header
    max_col = 18  # A-R columns
    
    # Set standard font and alignment for all cells
    for row in ws.iter_rows(min_row=1, max_row=max_row, min_col=1, max_col=max_col):
        for cell in row:
            cell.font = Font(name="Calibri", size=11)
            cell.alignment = Alignment(wrap_text=True, vertical="top")
    
    # Set column widths to match old working format exactly
    column_widths = {
        'A': 12,   # Job ID - hyperlinks
        'B': 60,   # Job description - wide for content
        'C': 30,   # Position title
        'D': 24,   # Location
        'E': 29,   # Job domain
        'F': 13,   # Match level
        'G': 21,   # Evaluation date
        'H': 16,   # Has domain gap
        'I': 36,   # Domain assessment
        'J': 36,   # No-go rationale
        'K': 36,   # Application narrative
        'L': 25,   # export_job_matches_log
        'M': 25,   # generate_cover_letters_log
        'N': 36,   # reviewer_feedback
        'O': 20,   # mailman_log
        'P': 25,   # process_feedback_log
        'Q': 25,   # reviewer_support_log
        'R': 15    # workflow_status
    }
    
    for col_letter, width in column_widths.items():
        ws.column_dimensions[col_letter].width = width
    
    # Set AutoFilter on entire data range
    data_range = f"A1:R{max_row}"
    ws.auto_filter.ref = data_range
    
    # Add conditional formatting for Match level column (F)
    apply_match_level_formatting(ws, max_row)
    
    # Freeze panes at B2 (freeze column A and row 1)
    ws.freeze_panes = ws['B2']

def apply_match_level_formatting(ws, max_row):
    """Apply color coding to the Match level column (F)"""
    for row in range(2, max_row + 1):
        cell = ws.cell(row=row, column=6)  # Column F
        cell_value = str(cell.value).strip().lower()
        
        if cell_value == 'good':
            cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")  # Light green
            cell.font = Font(name="Calibri", size=11, color="006100")  # Dark green text
        elif cell_value == 'low':
            cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")  # Light red
            cell.font = Font(name="Calibri", size=11, color="9C0006")  # Dark red text
        elif cell_value == 'moderate':
            cell.fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")  # Light yellow
            cell.font = Font(name="Calibri", size=11, color="9C6500")  # Dark yellow text
        elif cell_value == 'unknown':
            cell.fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")  # Light gray
            cell.font = Font(name="Calibri", size=11, color="808080")  # Gray text

def extract_matching_results(job_data):
    """Extract matching results from job data (legacy function)"""
    web_details = job_data.get("web_details", {})
    llama_eval = job_data.get("llama32_evaluation", {})
    
    return {
        'job_id': job_data.get('job_id'),
        'position_title': web_details.get('position_title', ''),
        'match_level': llama_eval.get('cv_to_role_match', ''),
        'domain_assessment': llama_eval.get('domain_knowledge_assessment', ''),
        'url': web_details.get('url', '')
    }

def export_job_matches(output_format='excel', output_file=None, job_ids=None, feedback_system=False, reviewer_name="xai"):
    """
    Export job matching results to CSV or Excel format.
    
    Args:
        output_format: Format to export to ('csv' or 'excel')
        output_file: Path to the output file. If None, a default name will be used.
        job_ids: List of specific job IDs to export. If None, all jobs with llama32_evaluation will be exported.
        feedback_system: Whether to export in feedback system format with A-R columns
        reviewer_name: Reviewer name for filename (default: xai)
        
    Returns:
        Path to the generated file
    """
    # Default to CSV if openpyxl not available and Excel requested
    if output_format.lower() == 'excel' and not OPENPYXL_AVAILABLE:
        print("Warning: openpyxl not available, falling back to CSV format")
        output_format = 'csv'
    
    # Ensure job_ids is a list
    if job_ids is not None and not isinstance(job_ids, list):
        job_ids = [job_ids]
    
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    
    # Prepare output file path
    if output_file is None:
        if feedback_system:
            output_dir = Path("output/excel")
            output_dir.mkdir(parents=True, exist_ok=True)
            output_file = output_dir / f"job_matches_{timestamp}.xlsx"
        else:
            output_file = f"job_export_{timestamp}.{output_format}"
    
    # Ensure .xlsx extension for Excel output
    if output_format.lower() == 'excel':
        if not str(output_file).lower().endswith('.xlsx'):
            output_file = str(output_file).rsplit('.', 1)[0] + '.xlsx'
    
    # Collect all job files
    job_files = list(Path(JOB_DATA_DIR).rglob("*.json"))
    processed_jobs = []
    
    print(f"Found {len(job_files)} job files to process...")
    
    # Process each job file
    for job_file in job_files:
        try:
            with open(job_file, 'r', encoding='utf-8') as f:
                job_data = json.load(f)
                
                # Get job_id from either legacy or beautiful format
                job_id = job_data.get('job_id') or job_data.get('job_metadata', {}).get('job_id')
                
                # Filter job IDs if specified
                if job_ids is not None and str(job_id) not in [str(jid) for jid in job_ids]:
                    continue
                
                # Check for evaluations in both legacy and beautiful formats
                has_evaluation = (
                    job_data.get('llama32_evaluation') or  # Legacy format
                    (job_data.get('evaluation_results', {}).get('cv_to_role_match'))  # Beautiful format
                )
                
                if has_evaluation:
                    processed_jobs.append(job_data)
                    
        except Exception as e:
            print(f"Warning: Could not process {job_file}: {e}")
            continue
    
    print(f"Processing {len(processed_jobs)} jobs with evaluations...")
    
    if not processed_jobs:
        print("No jobs with evaluations found to export.")
        return None
    
    # Export to the desired format
    try:
        if feedback_system:
            # JMFS feedback system format with A-R columns
            matches = []
            for job_data in processed_jobs:
                job_match = extract_job_data_for_feedback_system(job_data)
                logging_cols = initialize_logging_columns(timestamp)
                matches.append({**job_match, **logging_cols})
            
            export_to_excel_feedback_system(matches, output_file)
        else:
            # Legacy format
            matches = [extract_matching_results(job_data) for job_data in processed_jobs]
            if output_format.lower() == 'excel':
                export_to_excel(matches, output_file)
            else:
                export_to_csv(matches, output_file)
        
        print(f"Successfully exported {len(processed_jobs)} jobs to: {output_file}")
        return str(output_file)
        
    except Exception as e:
        print(f"Error during export: {e}")
        raise

if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="Export job matching results to CSV or Excel format.")
    parser.add_argument("--output-format", choices=["csv", "excel"], default="excel", 
                       help="Output format (csv or excel)")
    parser.add_argument("--output-file", help="Path to output file")
    parser.add_argument("--job-ids", nargs="*", type=int, help="List of job IDs to export")
    parser.add_argument("--feedback-system", action="store_true", 
                       help="Export in feedback system format with A-R columns")
    parser.add_argument("--reviewer-name", default="xai", 
                       help="Reviewer name for filename (default: xai)")
    
    args = parser.parse_args()
    
    try:
        output_path = export_job_matches(
            output_format=args.output_format,
            output_file=args.output_file,
            job_ids=args.job_ids,
            feedback_system=args.feedback_system,
            reviewer_name=args.reviewer_name
        )
        
        if output_path:
            print(f"Export completed successfully: {output_path}")
        else:
            print("Export failed - no data to export")
            sys.exit(1)
            
    except Exception as e:
        print(f"Export failed with error: {e}")
        sys.exit(1)