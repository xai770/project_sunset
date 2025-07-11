#!/usr/bin/env python3
"""
Export job matching results to CSV or Excel format.

This script reads job JSON files from the data/postings directory and exports
the matching results to a table format (CSV or Excel) for easier review and analysis.
It specifically focuses on extracting domain knowledge assessment information.
"""
import os
import sys
import json
import argparse
import glob
from pathlib import Path
import pandas as pd
from datetime import datetime
import warnings

# Try to import openpyxl for Excel formatting
try:
    import openpyxl
    from openpyxl.styles import PatternFill, Font, Alignment
    from openpyxl.styles.colors import Color
    from openpyxl.worksheet.dimensions import ColumnDimension
    from openpyxl.utils.cell import get_column_letter
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False
    warnings.warn("openpyxl not available. Excel formatting will be limited.")

from run_pipeline.config.paths import JOB_DATA_DIR

def get_standard_columns():
    """Define the standard column structure A-R for feedback system"""
    return {
        'A': 'URL',
        'B': 'Job description',
        'C': 'Position title',
        'D': 'Location',
        'E': 'Job domain',
        'F': 'Match level',
        'G': 'Evaluation date',
        'H': 'Has domain gap',
        'I': 'Domain assessment',
        'J': 'No-go rationale',
        'K': 'Application narrative',
        'L': 'export_job_matches_log',
        'M': 'generate_cover_letters_log',
        'N': 'reviewer_feedback',
        'O': 'mailman_log',
        'P': 'process_feedback_log',
        'Q': 'reviewer_support_log',
        'R': 'workflow_status'
    }

def extract_job_domain_from_data(job_data):
    """Extract job domain from job data"""
    # Try multiple fields for domain information
    web_details = job_data.get("web_details", {})
    search_details = job_data.get("search_details", {})
    
    # Check various possible domain fields
    domain = (
        job_data.get('domain') or 
        web_details.get('domain') or 
        search_details.get('JobCategory', [{}])[0].get('Name', '') if search_details.get('JobCategory') else ''
    )
    
    return domain

def determine_domain_gap(llama_eval):
    """Determine if there's a domain gap based on evaluation"""
    # Look for domain gap indicators in the evaluation
    domain_assessment = llama_eval.get('domain_knowledge_assessment', '').lower()
    no_go_rationale = llama_eval.get('no_go_rationale', '').lower()
    
    # Check for gap indicators
    gap_indicators = ['lacks', 'missing', 'no experience', 'gap', 'would take', 'years to acquire']
    has_gap = any(indicator in domain_assessment or indicator in no_go_rationale 
                  for indicator in gap_indicators)
    
    return 'Yes' if has_gap else 'No'

def extract_job_data_for_feedback_system(job_data):
    """Extract job data in the A-R column format for feedback system"""
    web_details = job_data.get("web_details", {})
    llama_eval = job_data.get("llama32_evaluation", {})
    search_details = job_data.get("search_details", {})
    
    # Get job_id directly from job data
    job_id = job_data.get('job_id')
    
    # Construct URL using job_id
    url = web_details.get('url', '')
    if not url and job_id:
        url = f"https://careers.db.com/professionals/search-roles/#/professional/job/{job_id}"
    
    # Extract location information
    location = ''
    if search_details.get('PositionLocation'):
        loc_data = search_details['PositionLocation'][0]
        city = loc_data.get('CityName', '')
        country = loc_data.get('CountryName', '')
        location = f"{city}, {country}".strip(', ')
    
    # Extract evaluation date
    eval_date = ''
    if llama_eval.get('processed_at'):
        try:
            eval_date = datetime.fromisoformat(llama_eval['processed_at'].replace('Z', '+00:00')).strftime('%Y-%m-%d')
        except:
            eval_date = llama_eval.get('processed_at', '')
    
    return {
        'URL': url,
        'Job description': web_details.get('concise_description', ''),
        'Position title': web_details.get('position_title', ''),
        'Location': location,
        'Job domain': extract_job_domain_from_data(job_data),
        'Match level': llama_eval.get('cv_to_role_match', ''),
        'Evaluation date': eval_date,
        'Has domain gap': determine_domain_gap(llama_eval),
        'Domain assessment': llama_eval.get('domain_knowledge_assessment', ''),
        'No-go rationale': llama_eval.get('No-go rationale', ''),
        'Application narrative': llama_eval.get('application_narrative', '')
    }

def initialize_logging_columns(timestamp):
    """Initialize the logging columns L-R with default values"""
    return {
        'export_job_matches_log': f"Exported at {timestamp}, v1.0, success",
        'generate_cover_letters_log': '',
        'reviewer_feedback': '',
        'mailman_log': '',
        'process_feedback_log': '',
        'reviewer_support_log': '',
        'workflow_status': 'Exported'
    }

def export_to_csv(matches, output_file):
    """Export matches to CSV format (legacy function)"""
    df = pd.DataFrame(matches)
    df.to_csv(output_file, index=False, quoting=1)  # QUOTE_ALL

def export_to_excel(matches, output_file):
    """Export matches to Excel format (legacy function)"""
    if not matches:
        raise ValueError("No matches to export.")
    
    first = matches[0]
    if isinstance(first, dict):
        header = first.keys()
        rows = [m.values() for m in matches]
    elif isinstance(first, list):
        header = [f"col_{i+1}" for i in range(len(first))]
        rows = matches
    else:
        raise TypeError("Matches must be a list of dicts or list of lists.")
    
    df = pd.DataFrame(rows, columns=header)
    df.to_excel(output_file, index=False)
    return output_file

def export_to_excel_feedback_system(matches, output_file):
    """Export matches to Excel with full JMFS feedback system formatting"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Job Matches"
    
    # Get column structure
    columns = get_standard_columns()
    headers = list(columns.values())
    
    # Write headers
    ws.append(headers)
    
    # Format header row
    for cell in ws[1]:
        cell.font = Font(bold=True, name="Calibri")
        cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center")
    
    # Write data rows
    for row_idx, match in enumerate(matches, start=2):
        row_data = [match.get(col_name, '') for col_name in columns.values()]
        ws.append(row_data)
        
        # Set hyperlink in column A (URL)
        url = match.get('URL', '')
        cell_a = ws.cell(row=row_idx, column=1)
        
        if url:
            # Extract job_id for display text
            job_id = ''
            try:
                # Try to get job_id from URL
                parts = url.rstrip('/').split('/')
                if parts and parts[-1].isdigit():
                    job_id = parts[-1]
            except:
                pass
            
            display_text = f"Job {job_id}" if job_id else "Job Link"
            cell_a.value = display_text
            cell_a.hyperlink = url
            cell_a.style = "Hyperlink"
        else:
            cell_a.value = ""
        
        # Set row height to 60pt for text wrapping
        ws.row_dimensions[row_idx].height = 60
    
    # Apply formatting
    apply_feedback_system_formatting(ws, len(matches))
    
    # Save workbook
    wb.save(output_file)
    print(f"Excel file exported with {len(matches)} jobs to: {output_file}")

def apply_feedback_system_formatting(ws, num_data_rows):
    """Apply comprehensive formatting to match the old working Excel format"""
    max_row = num_data_rows + 1  # +1 for header
    max_col = 18  # A-R columns
    
    # Set standard font and alignment for all cells
    for row in ws.iter_rows(min_row=1, max_row=max_row, min_col=1, max_col=max_col):
        for cell in row:
            cell.font = Font(name="Calibri", size=11)
            cell.alignment = Alignment(wrap_text=True, vertical="top")
    
    # Set column widths to match old working format exactly
    column_widths = {
        'A': 12,   # URL - hyperlinks
        'B': 120,# Job description - wide for content
        'C': 30,   # Position title
        'D': 24,   # Location
        'E': 29,   # Job domain
        'F': 13,   # Match level
        'G': 21,   # Evaluation date
        'H': 16,   # Has domain gap
        'I': 36,   # Domain assessment
        'J': 36,   # No-go rationale
        'K': 36,   # Application narrative
        'L': 25,   # export_job_matches_log
        'M': 25,   # generate_cover_letters_log
        'N': 36,   # reviewer_feedback
        'O': 20,   # mailman_log
        'P': 25,   # process_feedback_log
        'Q': 25,   # reviewer_support_log
        'R': 15    # workflow_status
    }
    
    for col_letter, width in column_widths.items():
        ws.column_dimensions[col_letter].width = width
    
    # Set AutoFilter on entire data range
    data_range = f"A1:R{max_row}"
    ws.auto_filter.ref = data_range
    
    # Add conditional formatting for Match level column (F)
    apply_match_level_formatting(ws, max_row)
    
    # Freeze panes at B2 (freeze column A and row 1)
    ws.freeze_panes = ws['B2']

def apply_match_level_formatting(ws, max_row):
    """Apply color coding to the Match level column (F)"""
    for row in range(2, max_row + 1):
        cell = ws.cell(row=row, column=6)  # Column F
        cell_value = str(cell.value).strip().lower()
        
        if cell_value == 'good':
            cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")  # Light green
            cell.font = Font(name="Calibri", size=11, color="006100")  # Dark green text
        elif cell_value == 'low':
            cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")  # Light red
            cell.font = Font(name="Calibri", size=11, color="9C0006")  # Dark red text
        elif cell_value == 'moderate':
            cell.fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")  # Light yellow
            cell.font = Font(name="Calibri", size=11, color="9C6500")  # Dark yellow text
        elif cell_value == 'unknown':
            cell.fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")  # Light gray
            cell.font = Font(name="Calibri", size=11, color="808080")  # Gray text

def extract_matching_results(job_data):
    """Extract matching results from job data (legacy function)"""
    web_details = job_data.get("web_details", {})
    llama_eval = job_data.get("llama32_evaluation", {})
    
    return {
        'job_id': job_data.get('job_id'),
        'position_title': web_details.get('position_title', ''),
        'match_level': llama_eval.get('cv_to_role_match', ''),
        'domain_assessment': llama_eval.get('domain_knowledge_assessment', ''),
        'url': web_details.get('url', '')
    }

def export_job_matches(output_format='excel', output_file=None, job_ids=None, feedback_system=False, reviewer_name="xai"):
    """
    Export job matching results to CSV or Excel format.
    
    Args:
        output_format: Format to export to ('csv' or 'excel')
        output_file: Path to the output file. If None, a default name will be used.
        job_ids: List of specific job IDs to export. If None, all jobs with llama32_evaluation will be exported.
        feedback_system: Whether to export in feedback system format with A-R columns
        reviewer_name: Reviewer name for filename (default: xai)
        
    Returns:
        Path to the generated file
    """
    # Default to CSV if openpyxl not available and Excel requested
    if output_format.lower() == 'excel' and not OPENPYXL_AVAILABLE:
        print("Warning: openpyxl not available, falling back to CSV format")
        output_format = 'csv'
    
    # Ensure job_ids is a list
    if job_ids is not None and not isinstance(job_ids, list):
        job_ids = [job_ids]
    
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    
    # Prepare output file path
    if output_file is None:
        if feedback_system:
            output_dir = Path("output/excel")
            output_dir.mkdir(parents=True, exist_ok=True)
            output_file = output_dir / f"job_matches_{timestamp}.xlsx"
        else:
            output_file = f"job_export_{timestamp}.{output_format}"
    
    # Ensure .xlsx extension for Excel output
    if output_format.lower() == 'excel':
        if not str(output_file).lower().endswith('.xlsx'):
            output_file = str(output_file).rsplit('.', 1)[0] + '.xlsx'
    
    # Collect all job files
    job_files = list(Path(JOB_DATA_DIR).rglob("*.json"))
    processed_jobs = []
    
    print(f"Found {len(job_files)} job files to process...")
    
    # Process each job file
    for job_file in job_files:
        try:
            with open(job_file, 'r', encoding='utf-8') as f:
                job_data = json.load(f)
                
                # Filter job IDs if specified
                if job_ids is not None and job_data.get('job_id') not in job_ids:
                    continue
                
                # Only include jobs with llama32_evaluation
                if job_data.get('llama32_evaluation'):
                    processed_jobs.append(job_data)
                    
        except Exception as e:
            print(f"Warning: Could not process {job_file}: {e}")
            continue
    
    print(f"Processing {len(processed_jobs)} jobs with evaluations...")
    
    if not processed_jobs:
        print("No jobs with evaluations found to export.")
        return None
    
    # Export to the desired format
    try:
        if feedback_system:
            # JMFS feedback system format with A-R columns
            matches = []
            for job_data in processed_jobs:
                job_match = extract_job_data_for_feedback_system(job_data)
                logging_cols = initialize_logging_columns(timestamp)
                matches.append({**job_match, **logging_cols})
            
            export_to_excel_feedback_system(matches, output_file)
        else:
            # Legacy format
            matches = [extract_matching_results(job_data) for job_data in processed_jobs]
            if output_format.lower() == 'excel':
                export_to_excel(matches, output_file)
            else:
                export_to_csv(matches, output_file)
        
        print(f"Successfully exported {len(processed_jobs)} jobs to: {output_file}")
        return str(output_file)
        
    except Exception as e:
        print(f"Error during export: {e}")
        raise

if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="Export job matching results to CSV or Excel format.")
    parser.add_argument("--output-format", choices=["csv", "excel"], default="excel", 
                       help="Output format (csv or excel)")
    parser.add_argument("--output-file", help="Path to output file")
    parser.add_argument("--job-ids", nargs="*", type=int, help="List of job IDs to export")
    parser.add_argument("--feedback-system", action="store_true", 
                       help="Export in feedback system format with A-R columns")
    parser.add_argument("--reviewer-name", default="xai", 
                       help="Reviewer name for filename (default: xai)")
    
    args = parser.parse_args()
    
    try:
        output_path = export_job_matches(
            output_format=args.output_format,
            output_file=args.output_file,
            job_ids=args.job_ids,
            feedback_system=args.feedback_system,
            reviewer_name=args.reviewer_name
        )
        
        if output_path:
            print(f"Export completed successfully: {output_path}")
        else:
            print("Export failed - no data to export")
            sys.exit(1)
            
    except Exception as e:
        print(f"Export failed with error: {e}")
        sys.exit(1)